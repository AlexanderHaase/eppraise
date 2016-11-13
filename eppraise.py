#!/usr/bin/python3
from ebaysdk.finding import Connection as Finding
from ebaysdk.exception import ConnectionError

import itertools
import functools
import openpyxl
import sys
import re
import yaml
import argparse
import logging
import json
import datetime
import collections
from flask import Flask, redirect, request, render_template, make_response
from sqlalchemy.ext.declarative import declarative_base, declared_attr, as_declarative
from sqlalchemy import Table, Column, Integer, String, DateTime, ForeignKey, Boolean, create_engine
from sqlalchemy.orm import relationship, sessionmaker
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm.exc import NoResultFound

logger = logging.getLogger( __name__ )

def consume(iterator, n = None):
    "Advance the iterator n-steps ahead. If n is none, consume entirely."
    # Use functions that consume iterators at C speed.
    if n is None:
        # feed the entire iterator into a zero-length deque
        collections.deque(iterator, maxlen=0)
    else:
        # advance to the empty slice starting at position n
        next(islice(iterator, n, n), None)

def apply( function, iterator ):
	consume( map( function, iterator ) )

def unique( iterator ):
	seenValues = set()
	while True:
		value = next( iterator )
		if value not in seenValues:
			seenValues.add( value )
			yield value

def scrub( keywords ):
	keys = re.sub( '(\s)', ' ', keywords ).split()
	filtered = map( lambda key: re.sub( '(\W)', '', key ), keys )
	return ' '.join( filter( None, filtered ) )

#
# Setup SQL Schema
#
@as_declarative()
class SQLBase( object ):
	'''Common properties for all sql objects'''
	@declared_attr
	def __tablename__(cls):
		return cls.__name__.lower()

	id = Column( Integer, primary_key = True, autoincrement = True )

	def dict( self, append = tuple(), exclude = ('text',) ):
		'''Serialize keys via reflection'''
		keys = itertools.chain( self.__table__.columns.keys(), append )
		keys = filter( lambda key: key not in exclude, keys )

		attrs = map( functools.partial( getattr, self ), dir( self ) )
		funcs = filter( lambda attr: hasattr( attr, '__serialize__' ) and attr.__serialize__, attrs )
 
		result = collections.OrderedDict( map( lambda key: (key, getattr( self, key )), keys ) )
		result.update( collections.OrderedDict( map( lambda func: (func.__name__, func()), funcs ) ) )
		return result

	def serialize( func ):
		func.__serialize__ = True
		return func


class JSONProps( object ):
	'''Mix-in for text<->json'''
	
	text = Column( String, nullable = False )

	@property
	def json( self ):
		if not hasattr( self, '__jsonCache__'):
			self.__jsonCache__ = json.loads( self.text )
		return self.__jsonCache__

	@json.setter
	def json( self, data ):
		self.text = json.dumps( data )
		self.__jsonCache__ = data # TODO: Deep copy/reload data....


associate_watch_item = Table( 'associate_watch_item', SQLBase.metadata, 
	Column( "watch_id", Integer, ForeignKey( "watch.id" ), primary_key = True ),
	Column( "item_id", Integer, ForeignKey( "item.id" ), primary_key = True )
)
	

class Watch( SQLBase ):
	'''Saved watch for items'''
	keywords = Column( String, nullable = False, unique = True )
	enabled = Column( Boolean, nullable = False, default = True )

	items = relationship("Item", back_populates="watches", secondary = associate_watch_item )
	queries = relationship("Query", back_populates="watch")

	@SQLBase.serialize
	def estimate( self ):
		'''Mean sold price'''
		(total, qty) = functools.reduce( (lambda accum, item: ( accum[ 0 ] + item.price(), accum[ 1 ] + 1.0 ) ), self.items, ( 0.0, 0.0 ) )
		return total / qty if qty > 0 else None
		
	@classmethod
	def queryAll( cls, context, connection ):
		activeWatches = context.session().query( cls ).filter( cls.enabled == True ).all()
		return map( functools.partial( Query.fromWatch, context, connection ), activeWatches )

	@classmethod
	def fromFile( cls, context, filePath, inputRange ):
		wb = openpyxl.load_workbook( filePath )
		sheet = wb.active
		return map( lambda cell: context.upsert( Watch, keywords = cell.value ), itertools.chain.from_iterable( sheet[ inputRange ] ) )


class Query( SQLBase, JSONProps ):
	'''Record of executing a query. Future-proofing our data!'''
	watch_id = Column( Integer, ForeignKey( Watch.id ), nullable = False )
	watch = relationship( Watch, back_populates = "queries" )
	retrieved = Column( DateTime, default = datetime.datetime.utcnow, nullable = False )
	keywords = Column( String, nullable = False )

	@classmethod
	def fromWatch( cls, context, connection, watch ):
		'''Create a query from a watch'''
		keywords = scrub( watch.keywords )
		result = connection.query( keywords )
		return context.upsert( cls, keywords = keywords, watch = watch, json = result.dict() )



class Item( SQLBase, JSONProps ):
	'''Record from ebay. We're watching completed items, so one per item is enough.'''
	ebayID = Column( String, unique = True, nullable = False )

	watches = relationship( Watch, back_populates = "items", secondary = associate_watch_item )

	@SQLBase.serialize
	def date( self ):
		return self.json[ 'listingInfo' ][ 'endTime' ]

	@SQLBase.serialize
	def url( self ):
		return self.json[ 'viewItemURL' ]

	@SQLBase.serialize
	def sold( self ):
		return self.json[ 'sellingStatus' ][ 'sellingState' ] == 'EndedWithSales'

	@SQLBase.serialize
	def price( self ):
		'''Fetch an iterator of sold prices'''
		return float( self.json[ 'sellingStatus' ][ 'currentPrice' ][ 'value' ] )

	
	@classmethod
	def fromQuery( cls, context, query ):
		'''Creates NEW objects from query'''
		items = query.json[ 'searchResult' ].get( 'item', tuple() )
		return map( lambda item: context.upsert( cls, watches = [ query.watch ], json = item, ebayID = item['itemId'] ), items )


class Database( object ):

	def __init__( self, dbURL = 'sqlite:///:memory:', base = SQLBase ):
		self.engine = create_engine( dbURL )
		self.sessionMaker = sessionmaker( self.engine )
		base.metadata.create_all( self.engine )

	def refresh( self ):
		self.engine = create_engine( self.dbURL )
		self.sessionMaker = sessionmaker( self.engine )
		base.metadata.create_all( self.engine )

	class SessionContext( object ):
		def __init__( self, db ):
			self.db = db

		def session( self ):
			return self.activeSession

		def __call__( self, func, key = 'context' ):
			'''decorator'''
			@functools.wraps( func )
			def wrapper( *args, **kwargs ):
				with self:
					kwargs[ key ] = self
					return func( *args, **kwargs )
			return wrapper

		def __enter__( self ):
			self.activeSession = self.db.sessionMaker()
			return self

		def __exit__( self, type, value, traceback ):
			if value:
				self.activeSession.rollback()
			else:
				self.activeSession.commit()
			self.activeSession.close()
			del self.activeSession

		def refresh( self ):
			self.activeSession.rollback()
			self.activeSession.close()
			self.activeSession = self.db.sessionMaker()

		@staticmethod
		def identifyingColumns( cls ):
			return filter( lambda column: column.unique or column.primary_key, cls.__table__.columns )

		@classmethod
		def queryArgs( this, cls, kwargs ):
			present = filter( lambda column: column.name in kwargs, this.identifyingColumns( cls ) )
			return map( lambda column: getattr( cls, column.name ) == kwargs[ column.name ], present )

		@staticmethod
		def updateKey( obj, key, value ):
			if key in obj.__class__.__table__.columns:
				setattr( obj, key, value )
			elif hasattr( obj, key ) and isinstance( getattr( obj, key ), list ):
				getattr( obj, key ).extend( value )
			else:
				setattr( obj, key, value )

		def upsert( self, cls, **kwargs ):
			try:
				queryArgs = tuple(self.queryArgs( cls, kwargs) )

				if len( queryArgs ) == 0:
					raise KeyError( queryArgs )

				obj = self.session().query( cls ).filter( *queryArgs ).one()
				logger.info( "Already exists: {} {}".format( obj.__class__.__name__, obj.dict() ) )
				apply( lambda item: self.updateKey( obj, item[ 0 ], item[ 1 ] ), kwargs.items() )

			except (NoResultFound, KeyError):
				obj = cls( **kwargs )
				self.session().add( obj )
				logger.info( "Added new item: {} {}".format( obj.__class__.__name__, obj.dict() ) )

			return obj
			

		def commitIfNew( self, obj ):
			try:
				self.activeSession.add( obj )
				self.activeSession.commit()
				logger.info( "Added new item: {} {}".format( obj.__class__.__name__, obj.dict() ) )
				return True

			except IntegrityError as e:
				self.activeSession.rollback()
				logger.info( "Already exists: {} {}\n{}".format( obj.__class__.__name__, obj.dict(), e ) )
				return False

	def context( self ):
		return self.SessionContext( self )

	
#
# Ebay connections
#

class Estimate( object ):
	'''
	Estimate for a single search result, focusing on sold items. Mostly
	focuses on dynamically extracting features from data, rather than
	statically compution/storing them.
	
	'''

	def __init__( self, keyword, result ):
		self.raw = result
		self.keyword = keyword
		self.items = result.dict()[ 'searchResult' ].get( 'item', tuple() )


	def sold( self ):
		'''Fetch an iterator of sold items'''
		return filter( (lambda item: item[ 'sellingStatus' ][ 'sellingState' ] == 'EndedWithSales' ), self.items )

	def prices( self ):
		'''Fetch an iterator of sold prices'''
		return map( (lambda item: item[ 'sellingStatus' ][ 'currentPrice' ][ 'value' ] ), self.sold() )

	def mean( self ):
		'''Mean sold price'''
		(total, qty) = functools.reduce( (lambda accum, price: ( accum[ 0 ] + float(price), accum[ 1 ] + 1.0 ) ), self.prices(), ( 0.0, 0.0 ) )
		return total / qty if qty > 0 else None
		

class Connection( object ):
	'''Syntatic sugar for interacting with the ebay sdk'''

	def __init__( self, **kwargs ):
		self.api = Finding( **kwargs )

	def query( self, item ):
		return self.api.execute( 'findCompletedItems', {'keywords': item, } )

	def estimate( self, item ):
		'''Create an estimate for the given item'''
		return Estimate( item, self.query( item ) )

	def estimateFile( self, file, inputRange, outputRange ):
		'''Proof of concept method for dumping this to/from a file'''
		wb = openpyxl.load_workbook( file )
		sheet = wb.active
		ioRange = zip( sheet[ inputRange ], sheet[ outputRange ] )
		
		def handleElement( ioElement ):
			keys = re.sub( '(\s)', ' ', ioElement[ 0 ][ 0 ].value ).split()
			filtered = map( lambda key: re.sub( '(\W)', '', key ), keys )
			key = ' '.join( filter( None, filtered ) )
			sys.stderr.write( key )
			est = self.estimate( key )
			mean = est.mean()
			sys.stderr.write( ': {}\n'.format( mean ) )
			ioElement[ 1 ][ 0 ].value = mean

		functools.reduce( lambda x,y: None, map( handleElement, ioRange ) )
		wb.save( file )


if __name__ == '__main__':
	logger = logging.getLogger()
	logger.setLevel(logging.DEBUG)

	# create console handler and set level to info
	handler = logging.StreamHandler()
	handler.setLevel( logging.DEBUG )
	handler.setFormatter( logging.Formatter( "%(asctime)s - %(levelname)s\t<%(name)s:%(lineno)d>: %(message)s" ) )
	logger.addHandler( handler )

	# Setup parser
	parser = argparse.ArgumentParser()
	parser.add_argument('-d','--database-url', default = 'sqlite:///eppraise.db', help = "Database connection url" )
	parser.add_argument( '-v', '--verbose', default = "WARNING", help = "Set logging level in DEBUG, INFO, WARNING, ERROR, CRITICAL" )
	subparsers = parser.add_subparsers( help = "command help", dest = "command" )

	xlsxParser = subparsers.add_parser( "xlsx", help = "Interact with spreadsheet" )
	xlsxParser.add_argument("spreadsheet", help = "Input spreadsheet" )
	xlsxParser.add_argument('-i', '--input-range', required = True, help = "Range of items in spreadsheet to estimate, one per cell" )
	xlsxParser.add_argument('-o', '--output-range', required = False, help = "Range for output estimates in sheet, same size as input range" )

	watchParser = subparsers.add_parser( "watch", help = "Create or modify a watch" )
	watchParser.add_argument( "watch", help = "Keywords to watch" )
	watchParser.add_argument( "--disable", action='store_true', help = "Disable specified watch" )
	watchParser.add_argument( "--estimate", action='store_true', help = "Provide an estimate based on database data" )

	queryParser = subparsers.add_parser( "update", help = "Update all active watches" )
	queryParser.add_argument('-c','--config', default = "./config.yaml", help = "Configuration for ebay API" )

	itemParser = subparsers.add_parser( "item" )

	webParser = subparsers.add_parser( "web" )
	webParser.add_argument( '-a', '--host', default = "0.0.0.0", help = "Host IP address for binding server" )
	webParser.add_argument( '-p', '--port', default = "5000", help = "Host port for binding server", type = int )

	#todo webParser..

	# parse args
	args = parser.parse_args();

	# setup logger
	logger.setLevel( getattr( logging, args.verbose ) )
	logger.debug( args )

	# connect to database
	logger.debug( "Connecting to database: '{}'...".format( args.database_url ) )
	db = Database( args.database_url )

	if args.command == 'xlsx':
		with db.context() as context:

			def updateWatch( inputCell, outputCell ):
				watch = context.upsert( Watch, keywords = inputCell.value )
				if outputCell:
					outputCell.value = watch.estimate()


			workbook = openpyxl.load_workbook( args.spreadsheet )
			sheet = workbook.active

			inputCells = itertools.chain.from_iterable( sheet[ args.input_range ] )

			if args.output_range:
				outputCells =  itertools.chain.from_iterable( sheet[ args.output_range ] )
			else:
				outputCells = itertools.repeat( None )

			consume( itertools.starmap( updateWatch, zip( inputCells, outputCells ) ) )

			workbook.save( args.spreadsheet )

			#map( lambda cell: context.upsert( Watch, keywords = cell.value ), itertools.chain.from_iterable( sheet[ inputRange ] ) )
			

			#watches = Watch.fromFile( context, args.spreadsheet, args.input_range )
			# TODO output range

	elif args.command == 'watch':
		with db.context() as context:
			try:
				watch = context.session().query( Watch ).filter( Watch.keywords == args.watch ).one()

			except NoResultFound:
				watch = Watch( keywords = args.watch )
				context.session().add( watch )

			watch.enabled = not args.disable

			print( watch.dict() )

	elif args.command == 'update':				

		# read config
		with open( args.config, 'r' ) as handle:
			config = yaml.safe_load( handle )

		# connect
		con = Connection( config_file = None, appid = config[ 'ebay' ][ 'id' ] )

		with db.context() as context:

			for query in Watch.queryAll( context, con ):
				
				# Commit and filter new items
				apply( context.session().expunge, itertools.filterfalse( Item.sold, Item.fromQuery( context, query ) ) )

	elif args.command == 'item':
		with db.context() as context:
			apply( sys.stdout.write, map( "{}\n".format, map( SQLBase.dict, context.session().query( Item ).all() ) ) )


	elif args.command == 'web':
		app = Flask( __name__ )

		def no_cache( func ):
			@functools.wraps( func )
			def wrapper( *args, **kwargs ):
				response = func( *args, **kwargs )
				if isinstance( response, str ):
					response = make_response( response, 200 )
				response.headers[ 'Cache-Control' ] = 'no-cache, no-store, must-revalidate'
				response.headers[ 'Pragma' ] = 'no-cache'
				return response

			return wrapper
				

		def serialize( iterator, status = 200 ):
			response = make_response( json.dumps( list( map( SQLBase.dict, iterator ) ) ), status )
			response.headers[ 'Content-Type' ] = 'appliction/json'
			return response

		@app.route( '/watch' )
		@no_cache
		@db.context()
		def watch( context ):
			return serialize( context.session().query( Watch ).all() )

		@app.route( '/watch/<int:watchId>/items' )
		@no_cache
		@db.context()
		def watchItems( watchId, context ):
			return serialize( context.session().query( Watch ).filter( Watch.id == watchId ).one().items )

		@app.route( '/' )
		def index():
			return render_template( 'index.html' )

		@app.route( '/jsonp/<name>' )
		@no_cache
		def jsonp( name ):
			return render_template( name )

		app.run( args.host, port = args.port, debug = True )



